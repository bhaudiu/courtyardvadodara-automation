#!/usr/bin/env python3
"""
Courtyard Vadodara — GRR extraction engine.

Parses the three daily source reports and returns a normalized dict of
GRR line items for the business date, plus a reconciliation check.

Sources:
  - Manager Flash  (PDF, E106)  -> rooms stats + revenue summary + reconciliation total
  - Trial Balance  (PDF, E100)  -> authoritative revenue by GL code
  - Daily Operations (XLSX)     -> F&B outlet net sales + covers (cross-check)
"""
import re
import sys
import json
import datetime as dt

import pdfplumber
import openpyxl


# ----------------------------- helpers --------------------------------------

def to_num(tok: str):
    """Parse an Indian-formatted number token like '15,14,250.25' or '- 720.34'."""
    if tok is None:
        return None
    s = str(tok).strip().replace(",", "")
    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:].strip()
    if s.endswith("-"):
        neg = True
        s = s[:-1].strip()
    if s in ("", "-", "."):
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


NUM_RE = re.compile(r"-?\s?\d[\d,]*\.?\d*-?")


def numbers_in(line: str):
    """Return list of parsed numbers appearing in a line, in order."""
    out = []
    for m in NUM_RE.findall(line):
        v = to_num(m)
        if v is not None:
            out.append(v)
    return out


# --------------------------- Manager Flash ----------------------------------

def parse_manager_flash(path: str) -> dict:
    """Extract the DAY (2026) column for the fields we need."""
    text = ""
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            text += (pg.extract_text() or "") + "\n"
    lines = [ln.rstrip() for ln in text.splitlines()]

    def day_value(label: str, exact_prefix=True):
        """First numeric token after an exact label at line start -> DAY 2026."""
        for ln in lines:
            s = ln.strip()
            if exact_prefix:
                if s.startswith(label) and (len(s) == len(label) or not s[len(label)].isalpha()):
                    rest = s[len(label):]
                    nums = numbers_in(rest)
                    if nums:
                        return nums[0]
            else:
                if label in s:
                    nums = numbers_in(s)
                    if nums:
                        return nums[0]
        return None

    # business date from header like "12-08-26"
    biz_date = None
    m = re.search(r"(\d{2})-(\d{2})-(\d{2})", text)
    if m:
        d, mo, y = m.groups()
        biz_date = dt.date(2000 + int(y), int(mo), int(d))

    return {
        "biz_date": biz_date,
        "total_rooms_available": day_value("Total Rooms in Hotel"),
        "rooms_occupied_net": day_value("Rooms Occupied minus Comp and House Use"),
        "complimentary_rooms": day_value("Complimentary Rooms"),
        "adr_net": day_value("ADR minus Comp and House"),
        "mf_room_revenue": day_value("Room Revenue"),
        "mf_fb_revenue": day_value("Food And Beverage Revenue"),
        "mf_other_revenue": day_value("Other Revenue"),
        "mf_total_revenue": day_value("Total Revenue"),
    }


# --------------------------- Trial Balance ----------------------------------

# GL code -> GRR bucket mapping (revenue section only)
def tb_bucket(code: str) -> str:
    c = int(code)
    if 10000 <= c <= 10999:
        return "room_revenue"
    if c == 20100:
        return "wine_shop"
    if c in (21123, 21124):
        return "meal_plan_812ce"
    if 21100 <= c <= 21299:
        return "ce_812"
    if 22100 <= c <= 22299:
        return "brubon"
    if 23100 <= c <= 23299:
        return "the_deck"
    if 24100 <= c <= 24299:
        return "ird"
    if 25100 <= c <= 25299:
        return "mbow"
    if 26100 <= c <= 26299:
        return "bgc"
    if 50100 <= c <= 50299:
        return "banquet"
    if 68100 <= c <= 69099:
        return "transportation"
    if c in (70011, 70101) or (70010 <= c <= 70012):
        return "guest_laundry"
    if c == 70014 or (70020 <= c <= 70029):
        return "spa"
    if c == 99202:
        return "round_off"
    return "other_misc"


def parse_trial_balance(path: str) -> dict:
    text = ""
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            text += (pg.extract_text() or "") + "\n"
    lines = text.splitlines()

    # Isolate the Revenue section: from the line "Revenue" to "Revenue Total"
    rev_lines, in_rev = [], False
    revenue_total = None
    for ln in lines:
        s = ln.strip()
        if s == "Revenue":
            in_rev = True
            continue
        if s.startswith("Revenue Total"):
            nums = numbers_in(s)
            revenue_total = nums[-1] if nums else None
            in_rev = False
            continue
        if in_rev:
            rev_lines.append(s)

    buckets = {}
    line_items = {}
    for s in rev_lines:
        m = re.match(r"^(\d{4,6})\s+(.*?)\s+(-?\s?[\d,]+\.\d{2}-?)\s*$", s)
        if not m:
            continue
        code, desc, amt = m.group(1), m.group(2).strip(), to_num(m.group(3))
        if amt is None:
            continue
        line_items[code] = {"desc": desc, "amount": amt}
        b = tb_bucket(code)
        buckets[b] = buckets.get(b, 0.0) + amt

    return {
        "revenue_total": revenue_total,
        "buckets": buckets,
        "line_items": line_items,
    }


# -------------------------- Daily Operations --------------------------------

DAILYOPS_MAP = {
    "BDQCC 812 CE": "ce_812",
    "BDQCC WINE SHOP": "wine_shop",
    "BDQCC ROOM SERVICE": "ird",
    "BDQCC SPA": "spa",
    "BDQCC BRUBON CAFE": "brubon",
    "BDQCC THE DECK": "the_deck",
    "BDQCC MBOW": "mbow",
    "BDQCC CONFERENCE": "banquet",
    "BDQCC LAUNDRY": "guest_laundry",
    "BDQCC BARODA GRILL C": "bgc",
}


def parse_daily_ops(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Reports"]
    outlets = {}          # bucket -> {"rev","covers","checks","avg"}
    total_sales = guests = checks = None
    biz_date = None
    in_rc_table = False   # only read outlets inside the "Revenue Center Name" table

    def numf(v):
        return float(v) if isinstance(v, (int, float)) else None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a == "Business Dates" and b:
            if isinstance(b, dt.datetime):
                biz_date = b.date()
            else:
                try:
                    d, mo, y = re.split(r"[/-]", str(b))
                    biz_date = dt.date(int(y), int(mo), int(d))
                except Exception:
                    pass
        if isinstance(a, str) and a.strip() == "Revenue Center Name":
            in_rc_table = True
            total_sales = ws.cell(r + 1, 2).value  # "Total" row directly under header
            guests = ws.cell(r + 1, 4).value       # col D = Guests
            checks = ws.cell(r + 1, 7).value       # col G = Checks
            continue
        if in_rc_table and isinstance(a, str) and a.strip().startswith("Order Type"):
            in_rc_table = False
        if in_rc_table and isinstance(a, str):
            key = a.strip()
            for name, bucket in DAILYOPS_MAP.items():
                if key.startswith(name):
                    outlets[bucket] = {
                        "rev": numf(ws.cell(r, 2).value),
                        "covers": numf(ws.cell(r, 4).value),
                        "avg": numf(ws.cell(r, 6).value),
                        "checks": numf(ws.cell(r, 7).value),
                    }
                    break
    return {
        "biz_date": biz_date,
        "outlets": outlets,
        "total_fb_sales": total_sales,
        "guests": guests,
        "checks": checks,
    }


# ------------------------------ assemble ------------------------------------

def build_grr_day(mf_path, tb_path, do_path) -> dict:
    mf = parse_manager_flash(mf_path)
    tb = parse_trial_balance(tb_path)
    do = parse_daily_ops(do_path)
    b = tb["buckets"]

    total_rooms = mf["total_rooms_available"]
    rooms_occ = mf["rooms_occupied_net"]
    comp = mf["complimentary_rooms"] or 0
    room_rev = b.get("room_revenue", mf["mf_room_revenue"])
    occupancy = (rooms_occ / total_rooms) if (rooms_occ and total_rooms) else None
    adr = mf["adr_net"]
    revpar = (room_rev / total_rooms) if (room_rev and total_rooms) else None

    other_block = {
        "transportation": b.get("transportation", 0.0),
        "guest_laundry": b.get("guest_laundry", 0.0),
        "wine_shop": b.get("wine_shop", 0.0),
        "spa": b.get("spa", 0.0),
        "other_misc": b.get("other_misc", 0.0),
        "round_off": b.get("round_off", 0.0),
    }
    other_total = sum(other_block.values())

    fb_block = {
        "ce_812": b.get("ce_812", 0.0),
        "meal_plan_812ce": b.get("meal_plan_812ce", 0.0),
        "ird": b.get("ird", 0.0),
        "brubon": b.get("brubon", 0.0),
        "the_deck": b.get("the_deck", 0.0),
        "bgc": b.get("bgc", 0.0),
        "banquet": b.get("banquet", 0.0),
        "mbow": b.get("mbow", 0.0),
    }
    fb_total = sum(fb_block.values())

    grand_total = room_rev + other_total + fb_total
    mf_total = mf["mf_total_revenue"]
    diff = grand_total - mf_total if mf_total else None

    biz_date = mf["biz_date"] or do["biz_date"]

    # ---- detailed CYV-GRR-style breakdown ----
    do_outlets = do["outlets"]
    total_covers = do["guests"]
    apc = (fb_total / total_covers) if total_covers else None

    OUTLET_ORDER = [
        ("ce_812", "812 CE"),
        ("meal_plan_812ce", "812 Meal Plan"),
        ("ird", "IRD"),
        ("brubon", "Brubon Café"),
        ("the_deck", "The Deck (Starbucks)"),
        ("bgc", "Baroda Grill (BGC)"),
        ("banquet", "Banquet"),
        ("mbow", "MBOW"),
    ]
    outlet_details = []
    for key, name in OUTLET_ORDER:
        rev = fb_block.get(key, 0.0)
        cov = do_outlets.get(key, {}).get("covers")
        outlet_details.append({
            "key": key, "name": name,
            "revenue": rev,
            "covers": cov,
            "avg_check": (rev / cov) if (cov and rev is not None) else None,
        })

    OOD_ORDER = [
        ("spa", "Spa"),
        ("transportation", "Transportation"),
        ("guest_laundry", "Guest Laundry"),
        ("wine_shop", "Wine Shop"),
        ("other_misc", "Others Misc."),
    ]
    ood = [{"key": k, "name": n, "revenue": other_block.get(k, 0.0)}
           for k, n in OOD_ORDER if other_block.get(k, 0.0)]

    operating_income = {
        "total_rooms": total_rooms,
        "occupied_rooms": rooms_occ,
        "occupancy_pct": occupancy,
        "adr": adr,
        "revpar": revpar,
        "total_covers": total_covers,
        "apc": apc,
        "room_revenue": room_rev,
        "fb_revenue": fb_total,
        "ood_revenue": other_total,
        "total_hotel_revenue": grand_total,
    }
    fb_details = {
        "total_rnl": fb_total - fb_block.get("banquet", 0.0) - fb_block.get("ird", 0.0),
        "banquet_other": fb_block.get("banquet", 0.0),
        "ird": fb_block.get("ird", 0.0),
    }

    return {
        "operating_income": operating_income,
        "fb_details": fb_details,
        "outlet_details": outlet_details,
        "ood": ood,
        "business_date": biz_date.isoformat() if biz_date else None,
        "rooms": {
            "total_available": total_rooms,
            "occupied_net": rooms_occ,
            "complimentary": comp,
            "occupancy_pct": occupancy,
            "adr": adr,
            "revpar": revpar,
            "room_revenue": room_rev,
        },
        "other_revenue": {**other_block, "total": other_total},
        "fb_revenue": {**fb_block, "total": fb_total},
        "totals": {
            "grand_total_revenue": grand_total,
            "grand_total_manager_flash": mf_total,
            "diff": diff,
        },
        "covers": {"guests": do["guests"], "checks": do["checks"], "fb_pos_total": do["total_fb_sales"]},
        "recon": {
            "tb_revenue_total": tb["revenue_total"],
            "mf_total_revenue": mf_total,
            "grr_grand_total": grand_total,
        },
        "_dailyops_outlets": do["outlets"],
    }


if __name__ == "__main__":
    mf_path, tb_path, do_path = sys.argv[1], sys.argv[2], sys.argv[3]
    data = build_grr_day(mf_path, tb_path, do_path)
    print(json.dumps(data, indent=2, default=str))
