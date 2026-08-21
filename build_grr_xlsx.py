#!/usr/bin/env python3
"""build_grr_xlsx.py — turn the reconstructed owner report in data.json into a
formatted GRR (Guest Room Revenue) owner-report workbook, grr.xlsx.

This is the same report the DBR tab shows on the site and the same structure the
owner receives: every line item across Today / Month-to-Date / Year-to-Date, each
vs This Year / Last Year / Budget / Variance. It is a point-in-time REPORT export
(final figures), so cells hold values, not formulas.

Run: python3 build_grr_xlsx.py [data.json] [grr.xlsx]
Called automatically by build_site.py after data.json is written.
"""
import sys, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CUR = '"₹"#,##,##0'     # ₹ with Indian digit grouping
NUM = '#,##,##0'
PCT = '0.0%'

INK = "1B2430"
ACCENT = "8A0F30"
HEADFILL = "EEF1F4"
SECFILL = "E7EBEF"
TOTFILL = "F2E6EA"
GOOD = "127A45"
BAD = "B3243F"
WHITE = "FFFFFF"

thin = Side(style="thin", color="C6CDD5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fmt_for(kind):
    if kind == "pct":
        return PCT
    if kind == "number":
        return NUM
    return CUR


def build(data_path="data.json", out_path="grr.xlsx"):
    data = json.load(open(data_path))
    sections = data.get("sections") or []
    prop = data.get("property", "Courtyard by Marriott Vadodara")
    latest = data.get("latest_date", "")
    gen = data.get("generated_at", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GRR Owner Report"
    ws.sheet_view.showGridLines = False

    # 12 columns: Line Item + Today(3) + MTD(4) + YTD(4)
    ncol = 12
    last_col = get_column_letter(ncol)

    def cell(r, c, v=None, *, bold=False, size=9, color=INK, fill=None,
             align="right", fmt=None, border=True, italic=False):
        cc = ws.cell(r, c)
        if v is not None:
            cc.value = v
        cc.font = Font(name="Arial", size=size, bold=bold, italic=italic, color=color)
        cc.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fill:
            cc.fill = PatternFill("solid", fgColor=fill)
        if fmt:
            cc.number_format = fmt
        if border:
            cc.border = BORDER
        return cc

    r = 1
    # ---- title block ----
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    cell(r, 1, prop, bold=True, size=15, color=ACCENT, align="left", border=False)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    cell(r, 1, "GRR — Daily Owner Report", bold=True, size=11, align="left", border=False)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    sub = f"Reporting day: {latest}"
    if gen:
        sub += f"     ·     Generated {gen} (UTC)     ·     Managed by NewcrestImage"
    cell(r, 1, sub, size=9, color="5A6672", align="left", border=False)
    r += 2

    # ---- group header ----
    grp_row = r
    cell(grp_row, 1, "Line Item", bold=True, size=9, color=WHITE, fill=INK, align="left")
    spans = [("Today", 2, 4), ("Month to Date", 5, 8), ("Year to Date", 9, 12)]
    for title, a, b in spans:
        ws.merge_cells(start_row=grp_row, start_column=a, end_row=grp_row, end_column=b)
        cell(grp_row, a, title, bold=True, size=9, color=WHITE, fill=ACCENT, align="center")
        for c in range(a + 1, b + 1):
            cell(grp_row, c, None, fill=ACCENT)
    r += 1

    # ---- sub header ----
    sub_row = r
    subs = ["", "Actual", "Budget", "Var %",
            "Actual", "Last Yr", "Budget", "Var %",
            "Actual", "Last Yr", "Budget", "Var %"]
    for i, s in enumerate(subs, start=1):
        cell(sub_row, i, s, bold=True, size=8.5, color=INK, fill=HEADFILL,
             align=("left" if i == 1 else "center"))
    r += 1

    # ---- body ----
    keymap = [
        ("today_ty", None), ("today_bud", None), ("today_var", "var"),
        ("mtd_ty", None), ("mtd_ly", None), ("mtd_bud", None), ("mtd_var", "var"),
        ("ytd_ty", None), ("ytd_ly", None), ("ytd_bud", None), ("ytd_var", "var"),
    ]
    for sec in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        cell(r, 1, sec.get("title", ""), bold=True, size=9, color=ACCENT,
             fill=SECFILL, align="left")
        for c in range(2, ncol + 1):
            cell(r, c, None, fill=SECFILL)
        r += 1
        for row in sec.get("rows", []):
            label = row.get("label", "")
            kind = row.get("kind", "currency")
            is_tot = "Total Hotel Revenue" in label
            fill = TOTFILL if is_tot else None
            cell(r, 1, label, bold=is_tot, size=9, color=INK, fill=fill, align="left")
            for i, (k, special) in enumerate(keymap, start=2):
                v = row.get(k)
                if special == "var":
                    col = GOOD if (v is not None and v >= 0) else BAD
                    cell(r, i, v, size=9, color=(col if v is not None else INK),
                         fill=fill, fmt=PCT, bold=is_tot)
                else:
                    cell(r, i, v, size=9, color=INK, fill=fill,
                         fmt=fmt_for(kind), bold=is_tot)
            r += 1

    # ---- widths & freeze ----
    ws.column_dimensions["A"].width = 34
    for c in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12.5
    ws.freeze_panes = f"B{sub_row + 1}"
    ws.print_title_rows = f"{grp_row}:{sub_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    wb.save(out_path)
    print(f"Wrote {out_path}: {len(sections)} sections, reporting day {latest}")


if __name__ == "__main__":
    dp = sys.argv[1] if len(sys.argv) > 1 else "data.json"
    op = sys.argv[2] if len(sys.argv) > 2 else "grr.xlsx"
    build(dp, op)
