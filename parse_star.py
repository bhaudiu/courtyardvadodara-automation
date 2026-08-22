#!/usr/bin/env python3
"""parse_star.py — parse an STR/STAR 'Glance' sheet into star.json.

Glance layout (My Property vs Competitive Set):
  Two tables — Performance and 'vs last year % Change' — each with
  Occupancy / ADR / RevPAR, three columns (My Prop, Comp Set, Index),
  four period rows (Current Month, YTD, Running 3 Month, Running 12 Month).
"""
import sys, json, re, datetime as dt
import openpyxl

# column letters -> 1-index base for the three metric blocks
COLS = {"occ": (7, 8, 9), "adr": (12, 13, 14), "revpar": (17, 18, 19)}  # G/H/I, L/M/N, Q/R/S
PERIODS = [("cm", 11), ("ytd", 13), ("r3", 15), ("r12", 17)]
CHG_PERIODS = [("cm", 25), ("ytd", 27), ("r3", 29), ("r12", 31)]


def numf(ws, r, c):
    v = ws.cell(r, c).value
    return float(v) if isinstance(v, (int, float)) else None


def block(ws, period_rows):
    out = {}
    for metric, (cmp_c, cs_c, idx_c) in COLS.items():
        out[metric] = {}
        for key, row in period_rows:
            out[metric][key] = {
                "mp": numf(ws, row, cmp_c),
                "cs": numf(ws, row, cs_c),
                "idx": numf(ws, row, idx_c),
            }
    return out


MONTHS = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
          "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

# Comp sheet monthly-trend rows (My Property / Comp Set / Index) per metric.
TREND_ROWS = {"occ": (21, 22, 23), "adr": (33, 34, 35), "revpar": (45, 46, 47)}
# YoY "% Chg" rows (My Property / Comp Set / Index) per metric.
TREND_CHG_ROWS = {"occ": (26, 27, 28), "adr": (38, 39, 40), "revpar": (50, 51, 52)}


def parse_comp_trend(wb):
    """Monthly time-series from the 'Comp' tab: My Property vs Competitive Set
    (and the index) for Occupancy / ADR / RevPAR, across the trailing months."""
    if "Comp" not in wb.sheetnames:
        return None
    ws = wb["Comp"]
    # month header on row 20; a year appears on row 19 at the first column of each year.
    cols, months, cur_year = [], [], None
    for c in range(3, 40):
        yv = ws.cell(19, c).value
        if isinstance(yv, (int, float)) and 2000 < yv < 2100:
            cur_year = int(yv)
        mon = ws.cell(20, c).value
        if mon in MONTHS and cur_year:
            cols.append(c)
            months.append(f"{cur_year}-{MONTHS[mon]:02d}")
        elif mon not in MONTHS and cols:
            break   # trailing YTD / running-average summary columns begin
    if not cols:
        return None

    def series(row):
        return [numf(ws, row, c) for c in cols]

    out = {"months": months}
    for key, (mp, cs, idx) in TREND_ROWS.items():
        cmp_r, ccs_r, cidx_r = TREND_CHG_ROWS[key]
        out[key] = {"mp": series(mp), "cs": series(cs), "idx": series(idx),
                    "mp_chg": series(cmp_r), "cs_chg": series(ccs_r), "idx_chg": series(cidx_r)}
    return out


def parse_glance(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Glance"]
    b4 = str(ws.cell(4, 2).value or "")
    m = re.search(r"For the Month of:\s*([A-Za-z]+ \d{4})", b4)
    period_label = m.group(1) if m else None
    m2 = re.search(r"Date Created:\s*([A-Za-z]+ \d{1,2}, \d{4})", b4)
    created = m2.group(1) if m2 else None
    prop = str(ws.cell(2, 2).value or "").split("  ")[0].strip()
    return {
        "period_label": period_label,
        "created": created,
        "property": prop,
        "perf": block(ws, PERIODS),
        "change": block(ws, CHG_PERIODS),
        "trend": parse_comp_trend(wb),
    }


if __name__ == "__main__":
    path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "star.json"
    monthly = parse_glance(path)
    doc = {"monthly": monthly, "weekly": None}
    with open(out_path, "w") as f:
        json.dump(doc, f, separators=(",", ":"), default=str)
    print(f"star.json: {monthly['period_label']} (created {monthly['created']})")
    print("occ CM:", monthly["perf"]["occ"]["cm"])
    print("adr CM:", monthly["perf"]["adr"]["cm"])
    print("revpar CM:", monthly["perf"]["revpar"]["cm"])
    print("occ CM %chg:", monthly["change"]["occ"]["cm"])
